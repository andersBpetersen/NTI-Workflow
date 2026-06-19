"""Parse LifeCycleDefinitionTransitions from Vault configuration Excel exports."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

TRANSITIONS_SHEET = "LifeCycleDefinitionTransitions"
STATES_SHEET = "LifeCycleDefinitionStates"

REQUIRED_TRANSITION_COLUMNS = {
    "lifecycle_definition": "LifeCycleDefinition",
    "from_state": "From State",
    "to_state": "To State",
    "security": "Security",
}

OPTIONAL_TRANSITION_COLUMNS = {
    "transition_id": "Id",
    "custom_job": "Custom JobTypes",
}

REQUIRED_STATE_COLUMNS = {
    "lifecycle_definition": "LifeCycleDefinition",
    "state": "State DisplayName",
    "security": "State Security",
}

OPTIONAL_STATE_COLUMNS = {
    "state_id": "Id",
}


@dataclass
class ParseResult:
    lifecycle_definitions: list[str] = field(default_factory=list)
    roles: list[str] = field(default_factory=list)
    nodes: list[dict[str, Any]] = field(default_factory=list)
    edges: list[dict[str, Any]] = field(default_factory=list)
    state_definitions: list[dict[str, Any]] = field(default_factory=list)
    meta: dict[str, Any] = field(default_factory=dict)


class TransitionParseError(Exception):
    """Raised when the Excel file cannot be parsed."""


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_sheet(workbook, sheet_name: str) -> Worksheet | None:
    for name in workbook.sheetnames:
        if name.strip().lower() == sheet_name.lower():
            return workbook[name]
    return None


def _require_sheet(workbook, sheet_name: str) -> Worksheet:
    sheet = _find_sheet(workbook, sheet_name)
    if sheet is None:
        raise TransitionParseError(
            "Excel-filen mangler arket LifeCycleDefinitionTransitions."
        )
    return sheet


def _find_header_column(sheet: Worksheet, header_row: int, header_text: str) -> int | None:
    last_col = sheet.max_column or 1
    for col_idx in range(1, last_col + 1):
        if _cell_text(sheet.cell(header_row, col_idx).value) == header_text:
            return col_idx - 1
    return None


def _find_header_row(sheet: Worksheet, required_headers: dict[str, str]) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, 26):
        column_map: dict[str, int] = {}
        missing: list[str] = []

        for field_name, header_text in required_headers.items():
            col_idx = _find_header_column(sheet, row_idx, header_text)
            if col_idx is None:
                missing.append(header_text)
            else:
                column_map[field_name] = col_idx

        if not missing:
            return row_idx, column_map

    missing_list = ", ".join(required_headers.values())
    raise TransitionParseError(
        f"Kunne ikke finde header-rækken i arket '{sheet.title}'. "
        f"Forventede kolonner: {missing_list}."
    )


def _get_cell(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return _cell_text(row[index])


def _make_node_id(lifecycle: str, state: str, multiple_lifecycles: bool) -> str:
    if multiple_lifecycles and lifecycle:
        return f"{lifecycle}::{state}"
    return state


def _add_roles_from_security(security: str, roles: set[str]) -> None:
    if not security:
        return

    for part in security.split(";"):
        chunk = part.strip()
        if not chunk:
            continue
        role = chunk.split(":", 1)[0].strip()
        if role:
            roles.add(role)


def _parse_states_sheet(sheet: Worksheet, roles: set[str]) -> list[dict[str, Any]]:
    try:
        header_row, column_map = _find_header_row(sheet, REQUIRED_STATE_COLUMNS)
    except TransitionParseError:
        return []

    optional_map: dict[str, int] = {}
    for field_name, header_text in OPTIONAL_STATE_COLUMNS.items():
        col_idx = _find_header_column(sheet, header_row, header_text)
        if col_idx is not None:
            optional_map[field_name] = col_idx

    states: list[dict[str, Any]] = []
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        lifecycle = _get_cell(row, column_map["lifecycle_definition"])
        state = _get_cell(row, column_map["state"])
        security = _get_cell(row, column_map["security"])
        state_id = _get_cell(row, optional_map.get("state_id"))

        if not lifecycle or not state:
            continue

        _add_roles_from_security(security, roles)
        states.append(
            {
                "id": state_id or f"state-{row_idx}",
                "lifecycleDefinition": lifecycle,
                "state": state,
                "security": security,
                "rowIndex": row_idx,
            }
        )

    return states


def parse_transitions_excel(file_bytes: bytes) -> ParseResult:
    """Parse transitions from an Excel workbook and return graph elements."""
    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise TransitionParseError(
            "Filen kunne ikke læses som en Excel-fil. Vælg en gyldig Vault Excel-eksport."
        ) from exc

    try:
        sheet = _require_sheet(workbook, TRANSITIONS_SHEET)
        header_row, column_map = _find_header_row(sheet, REQUIRED_TRANSITION_COLUMNS)

        optional_map: dict[str, int] = {}
        for field_name, header_text in OPTIONAL_TRANSITION_COLUMNS.items():
            col_idx = _find_header_column(sheet, header_row, header_text)
            if col_idx is not None:
                optional_map[field_name] = col_idx

        rows: list[dict[str, str]] = []
        warnings: list[str] = []
        roles: set[str] = set()

        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            lifecycle = _get_cell(row, column_map["lifecycle_definition"])
            from_state = _get_cell(row, column_map["from_state"])
            to_state = _get_cell(row, column_map["to_state"])
            security = _get_cell(row, column_map["security"])
            transition_id = _get_cell(row, optional_map.get("transition_id"))
            custom_job = _get_cell(row, optional_map.get("custom_job"))

            if not lifecycle or not from_state or not to_state:
                warnings.append(
                    f"Række {row_idx} sprunget over: mangler LifeCycleDefinition, From State eller To State."
                )
                continue

            _add_roles_from_security(security, roles)
            rows.append(
                {
                    "lifecycle_definition": lifecycle,
                    "from_state": from_state,
                    "to_state": to_state,
                    "security": security,
                    "transition_id": transition_id or str(row_idx),
                    "custom_job": custom_job,
                    "row_index": str(row_idx),
                }
            )

        if not rows:
            raise TransitionParseError(
                f"Ingen gyldige transitions fundet i arket '{TRANSITIONS_SHEET}'."
            )

        lifecycle_definitions = sorted({row["lifecycle_definition"] for row in rows})
        multiple_lifecycles = len(lifecycle_definitions) > 1

        states_sheet = _find_sheet(workbook, STATES_SHEET)
        state_definitions: list[dict[str, Any]] = []
        if states_sheet:
            state_definitions = _parse_states_sheet(states_sheet, roles)
            if not state_definitions:
                warnings.append(
                    f"Arket '{STATES_SHEET}' blev fundet men kunne ikke parses "
                    "(manglende kolonner eller ingen gyldige rækker)."
                )

        if "Everyone" not in roles:
            roles.add("Everyone")

        node_map: dict[str, dict[str, Any]] = {}
        edges: list[dict[str, Any]] = []

        for index, row in enumerate(rows, start=1):
            lifecycle = row["lifecycle_definition"]

            for state in (row["from_state"], row["to_state"]):
                node_id = _make_node_id(lifecycle, state, multiple_lifecycles)
                if node_id not in node_map:
                    node_map[node_id] = {
                        "id": node_id,
                        "label": state,
                        "lifecycleDefinition": lifecycle,
                    }

            source_id = _make_node_id(lifecycle, row["from_state"], multiple_lifecycles)
            target_id = _make_node_id(lifecycle, row["to_state"], multiple_lifecycles)

            edges.append(
                {
                    "id": row["transition_id"] or f"transition-{index}",
                    "source": source_id,
                    "target": target_id,
                    "label": row["custom_job"] or "",
                    "lifecycleDefinition": lifecycle,
                    "fromState": row["from_state"],
                    "toState": row["to_state"],
                    "security": row["security"],
                    "customJob": row["custom_job"],
                    "rowIndex": int(row["row_index"]),
                }
            )

        return ParseResult(
            lifecycle_definitions=lifecycle_definitions,
            roles=sorted(roles, key=str.lower),
            nodes=list(node_map.values()),
            edges=edges,
            state_definitions=state_definitions,
            meta={
                "sheetName": TRANSITIONS_SHEET,
                "statesSheetName": STATES_SHEET if states_sheet else None,
                "headerRow": header_row,
                "columnsDetected": {
                    **column_map,
                    **{f"optional_{key}": value for key, value in optional_map.items()},
                },
                "transitionCount": len(edges),
                "nodeCount": len(node_map),
                "stateDefinitionCount": len(state_definitions),
                "warnings": warnings,
            },
        )
    finally:
        workbook.close()


def parse_result_to_dict(result: ParseResult) -> dict[str, Any]:
    return {
        "lifecycleDefinitions": result.lifecycle_definitions,
        "roles": result.roles,
        "nodes": result.nodes,
        "edges": result.edges,
        "stateDefinitions": result.state_definitions,
        "meta": result.meta,
    }
