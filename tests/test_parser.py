"""Tests for lifecycle transition parser."""

import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.parser import TransitionParseError, parse_result_to_dict, parse_transitions_excel

SAMPLE_FILE = Path(__file__).resolve().parent.parent / "samples" / "sample-lifecycle.xlsx"


@pytest.fixture(scope="module", autouse=True)
def ensure_sample_file() -> None:
    if not SAMPLE_FILE.exists():
        from scripts.create_sample_excel import main

        main()


def test_parse_sample_excel() -> None:
    data = parse_transitions_excel(SAMPLE_FILE.read_bytes())
    payload = parse_result_to_dict(data)

    assert payload["meta"]["transitionCount"] == 5
    assert payload["meta"]["nodeCount"] == 4
    assert payload["meta"]["stateDefinitionCount"] == 4
    assert payload["lifecycleDefinitions"] == ["Basic Release Process"]
    assert "Engineer" in payload["roles"]
    assert "Everyone" in payload["roles"]

    node_labels = {node["label"] for node in payload["nodes"]}
    assert node_labels == {"Work In Progress", "For Review", "Released", "Obsolete"}

    first_edge = payload["edges"][0]
    assert "security" in first_edge
    assert "customJob" in first_edge
    assert first_edge["fromState"] == "Work In Progress"


def test_invalid_xlsx_content_raises() -> None:
    with pytest.raises(TransitionParseError, match="Filen kunne ikke læses som en Excel-fil"):
        parse_transitions_excel(b"not a real excel file")


def test_missing_sheet_raises() -> None:
    workbook = Workbook()
    workbook.active.title = "WrongSheet"
    buffer = io.BytesIO()
    workbook.save(buffer)

    with pytest.raises(TransitionParseError, match="LifeCycleDefinitionTransitions"):
        parse_transitions_excel(buffer.getvalue())


def test_missing_required_columns_raises() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LifeCycleDefinitionTransitions"
    sheet.append(["LifeCycleDefinition", "From State", "To State"])
    sheet.append(["LC1", "A", "B"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    with pytest.raises(TransitionParseError, match="Security"):
        parse_transitions_excel(buffer.getvalue())


def test_transitions_without_states_sheet() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LifeCycleDefinitionTransitions"
    sheet.append(
        ["Id", "LifeCycleDefinition", "From State", "To State", "Security"]
    )
    sheet.append(["1", "LC1", "Draft", "Released", "Everyone:Allow"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    payload = parse_result_to_dict(parse_transitions_excel(buffer.getvalue()))
    assert payload["meta"]["stateDefinitionCount"] == 0
    assert payload["meta"]["transitionCount"] == 1


def test_states_sheet_with_invalid_columns_does_not_fail() -> None:
    workbook = Workbook()
    transitions = workbook.active
    transitions.title = "LifeCycleDefinitionTransitions"
    transitions.append(
        ["Id", "LifeCycleDefinition", "From State", "To State", "Security"]
    )
    transitions.append(["1", "LC1", "Draft", "Released", "Everyone:Allow"])

    states = workbook.create_sheet("LifeCycleDefinitionStates")
    states.append(["WrongColumn", "AnotherColumn"])
    states.append(["x", "y"])

    buffer = io.BytesIO()
    workbook.save(buffer)

    payload = parse_result_to_dict(parse_transitions_excel(buffer.getvalue()))
    assert payload["meta"]["transitionCount"] == 1
    assert payload["meta"]["stateDefinitionCount"] == 0
    assert any("LifeCycleDefinitionStates" in w for w in payload["meta"]["warnings"])


def test_no_valid_transitions_raises() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LifeCycleDefinitionTransitions"
    sheet.append(
        ["Id", "LifeCycleDefinition", "From State", "To State", "Security"]
    )
    sheet.append(["1", "", "A", "B", "Everyone:Allow"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    with pytest.raises(TransitionParseError, match="Ingen gyldige transitions"):
        parse_transitions_excel(buffer.getvalue())
